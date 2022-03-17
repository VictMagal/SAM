import sys
from PyQt5 import uic, QtWidgets
import openpyxl

class load_arquivos:
        def __init__(self):
            self.arquivo_fornecedor = []
            self.list_fornecedor = []
            self.wb_new = openpyxl.Workbook()
            self.ws_new = self.wb_new.active
            
            
        def ler_arquivo_fornecedor(self):
            
            fornecedor = QtWidgets.QFileDialog.getOpenFileNames()[0]
            self.list_fornecedor = fornecedor
            
            print('Load file fornecedor complete!')
            print('list_fornecedor:', self.list_fornecedor)

        def juntar_arquivos(self):
            k = 0
            ultima_linha_ws_new = 1
            
            while k<= len(self.list_fornecedor)-1:
                self.wb = openpyxl.load_workbook(self.list_fornecedor[k])
                self.ws = self.wb ['Worksheet']
                
                is_data = True
                count_row_ws = 1
                while is_data:
                    count_row_ws += 1
                    data =  self.ws.cell (row = count_row_ws, column = 1).value
                    if data == None:
                        is_data = False
                count_row_ws -=1
                
                is_data = True
                count_col_ws = 1
                while is_data:
                    count_col_ws += 1
                    data =  self.ws.cell (row = 1, column = count_col_ws).value
                    if data == None:
                        is_data = False
                count_col_ws -=1
                
                for j in range(count_col_ws):
                    for i in range(count_row_ws):
                        if k ==0:
                            i-=1
                        self.ws_new.cell(row=ultima_linha_ws_new+i+1, column= j+1).value = self.ws.cell(row=i+2,column=j+1).value
                ultima_linha_ws_new += count_row_ws-1
                k+=1
            
            load_arquivos.excel_atualizador(self)
            

        def excel_atualizador(self):
            cliente = tela.lineEdit.text()
            banco = tela.lineEdit_2.text()
            
            wb_twm = openpyxl.load_workbook('./arquivo_twm/arquivo_'+cliente.upper()+'.xlsx')
            wb_atualizador = openpyxl.load_workbook('./arquivo_cliente/Modelo_Campo_Customizado.xlsx')
            
            ws_atualizador_dados = wb_atualizador ['Dados Cliente']
            ws_atualizador_fatura = wb_atualizador ['Fatura']
            ws_twm = wb_twm ['Planilha1']
            
            ws_atualizador_dados.cell(row = 2, column = 1).value = banco
            
            is_data = True
            count_row_ws = 1
            while is_data:
                count_row_ws += 1
                data =  ws_twm.cell (row = count_row_ws, column = 1).value
                if data == None:
                    is_data = False
            count_row_ws -=1

            i = 2
            while i <= count_row_ws:
                ws_atualizador_fatura.cell(row = i, column = 1).value = ws_twm.cell (row = i, column = 1).value
                ws_atualizador_fatura.cell(row = i, column = 2).value = 'Saneado'
                ws_atualizador_fatura.cell(row = i, column = 3).value = 'Sim'
             
                i += 1
            
            self.wb_new.save('_____CONSOLIDADO_COMPLETO'+banco+'.xlsx')
            print('Complete!!')
            
            wb_atualizador.save('__Modelo_Campo_Customizado_'+banco+'.xlsx')
            print('__Modelo_Campo_Customizado_'+banco+'.xlsx complete!!!')

app = QtWidgets.QApplication([])
tela = uic.loadUi("./assets/Interface_SAM.ui")

tela.show()

load_active = load_arquivos()
tela.pushButton.clicked.connect(load_active.ler_arquivo_fornecedor)
tela.pushButton_4.clicked.connect(load_active.juntar_arquivos)

sys.exit(app.exec_())